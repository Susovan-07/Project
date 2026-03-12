import os, re, sys, glob, math, difflib, unicodedata, warnings, json, urllib.request
from collections import Counter

import numpy as np
import pandas as pd
import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.neighbors import NearestNeighbors
from sklearn.cluster import AgglomerativeClustering
from sklearn.decomposition import TruncatedSVD
from sklearn.preprocessing import normalize
from scipy.sparse import hstack

warnings.filterwarnings("ignore")

# Optional packages — graceful upgrade if installed 
try:
    from sentence_transformers import SentenceTransformer
    _HAS_ST = True
except ImportError:
    _HAS_ST = False

try:
    import faiss as _faiss
    _HAS_FAISS = True
except ImportError:
    _HAS_FAISS = False



# CONFIGURATION


CFG = {
    # BM25
    "bm25_k1"            : 1.5,    # TF saturation
    "bm25_b"             : 0.75,   # Length normalization
    "bm25_candidates"    : 50,     # Stage-1 candidates per post

    # Semantic embedding (LSA / Sentence-Transformer)
    "embed_dims"         : 200,    # LSA dimensions (ignored when ST available)
    "st_model"           : "paraphrase-multilingual-MiniLM-L12-v2",
    "st_batch"           : 64,

    # Similarity thresholds
    "bm25_threshold"     : 0.20,   # min BM25 (normalized) to pass Stage 1
    "semantic_threshold" : 0.55,   # min semantic cosine for final output
    "tfidf_threshold"    : 0.15,   # TF-IDF comparison
    "fuzzy_threshold"    : 0.30,   # Fuzzy baseline

    # Output
    "top_n"              : 5,      # matches per post
    "fuzzy_sample"       : 150,    # records for fuzzy comparison

    # Clustering
    "cluster_size"       : 10,     # target posts per cluster

    # LLM titles
    "llm_model"          : "claude-sonnet-4-20250514",
    "llm_max_clusters"   : 40,     # clusters sent to LLM (rest get keyword titles)
    "anthropic_api_key"  : "",     # or set ANTHROPIC_API_KEY env var
}



# STEP 0: FILE RESOLVER


def resolve_input_file() -> str:
    if len(sys.argv) > 1:
        p = sys.argv[1]
        if os.path.exists(p): return p
        raise FileNotFoundError(f"Not found: {p}")
    for d in [os.getcwd(), os.path.dirname(os.path.abspath(__file__))]:
        hits = glob.glob(os.path.join(d, "*.xlsx"))
        if hits:
            print(f"  Auto-detected: {hits[0]}")
            return hits[0]
    fb = "/mnt/user-data/uploads/MATRIX_Hierarchy_2026-03-09.xlsx"
    if os.path.exists(fb): return fb
    raise FileNotFoundError(
        "\n[ERROR] No .xlsx file found.\n"
        "  Place the file next to this script OR run:\n"
        "  python3 post_similarity_matcher.py path/to/file.xlsx\n"
    )



# STEP 1: DATA LOADING


def load_all_sheets(filepath: str) -> pd.DataFrame:
    """
    Memory-efficient loading via openpyxl read_only=True.
    Sheet structure: Row1=title(skip), Row2=headers, Row3+=data.
    Column mapping: Content→post_snippet, Author→username.
    """
    print(f"\n[STEP 1] Loading: {filepath}")
    wb, records = openpyxl.load_workbook(filepath, read_only=True), []
    for sheet in wb.sheetnames:
        ws   = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        hdrs = [str(h).strip() if h else "" for h in rows[1]]
        for row in rows[2:]:
            if not any(row): continue
            rec = dict(zip(hdrs, row))
            txt = rec.get("Content", "") or ""
            if not str(txt).strip(): continue
            records.append({
                "id"           : len(records) + 1,
                "sheet"        : sheet,
                "district"     : str(rec.get("District", "") or "").strip(),
                "platform"     : str(rec.get("Platform", "") or "").strip(),
                "username"     : str(rec.get("Author",   "") or "").strip(),
                "post_snippet" : str(txt).strip(),
                "date"         : str(rec.get("Date",     "") or "").strip(),
            })
    df = pd.DataFrame(records)
    print(f"  Loaded {len(df):,} records across {len(wb.sheetnames)} sheets")
    return df



# STEP 2: MULTILINGUAL PREPROCESSING


HINDI_SW  = {"है","हैं","में","की","के","का","को","से","और","पर","यह","वह",
              "इस","उस","एक","था","थी","थे","हो","कि","जो","भी","नहीं","तो",
              "कर","रहा","रही","रहे","गया","गई","दे","दी","किया","लिया","ने",
              "हुए","हुई","हुआ","अपने","अपनी","अपना","अब","तब","जब","तक"}
ROMAN_SW  = {"ka","ki","ke","ko","se","me","mai","hai","hain","aur","par","ne",
              "yeh","woh","ek","bhi","nahi","to","jo","kiya","gaya","hua","hue",
              "raha","rahi","the","a","an","in","on","at","by","for","with",
              "from","is","was","are","were","be","been","have","has","had","do",
              "does","did","will","would","could","should","it","its","this",
              "that","as","of","or","and","but","not","no","so","up","out"}
ALL_SW = HINDI_SW | ROMAN_SW

def preprocess(text: str) -> str:
    """
    Full multilingual pipeline:
      Unicode NFC → remove URLs/emails → hashtag→word → remove @mentions →
      remove emojis → remove digits → remove punctuation (keep Devanagari) →
      lowercase Latin only → stopword removal → drop len≤1 tokens
    """
    if not text or not isinstance(text, str): return ""
    t = unicodedata.normalize("NFC", text)
    t = re.sub(r"http\S+|www\.\S+", " ", t)
    t = re.sub(r"\S+@\S+",          " ", t)
    t = re.sub(r"#(\w+)",            r"\1", t)
    t = re.sub(r"@\w+",              " ", t)
    t = re.sub(r"[\U00010000-\U0010FFFF]", " ", t)
    t = re.sub(r"[\u2000-\u206F\u2E00-\u2E7F]", " ", t)
    t = re.sub(r"\d+",               " ", t)
    t = re.sub(r"[^\w\s\u0900-\u097F]", " ", t)
    tokens = []
    for w in t.split():
        tokens.append(w if any("\u0900"<=c<="\u097F" for c in w) else w.lower())
    tokens = [tk for tk in tokens if tk not in ALL_SW and len(tk) > 1]
    return " ".join(tokens).strip()

def preprocess_df(df: pd.DataFrame) -> pd.DataFrame:
    print("\n[STEP 2] Preprocessing multilingual text...")
    df = df.copy()
    before = len(df)
    df = df.drop_duplicates(subset=["post_snippet"], keep="first").reset_index(drop=True)
    print(f"  Removed {before-len(df)} exact duplicates → {len(df):,} remain")
    df["post_snippet"]   = df["post_snippet"].fillna("").astype(str)
    df["processed_text"] = df["post_snippet"].apply(preprocess)
    df = df[df["processed_text"].str.len() > 5].reset_index(drop=True)
    df["id"] = range(1, len(df)+1)
    lens = df["post_snippet"].str.len()
    print(f"  Post length: min={lens.min()}, max={lens.max()}, avg={lens.mean():.0f} chars")
    print(f"  Ready: {len(df):,} records")
    return df



# STAGE 1 — BM25 FROM SCRATCH


class BM25:
    """
    BM25 (Best Match 25) — implemented fully from scratch.

    Formula per term t in document d:
        score(t,d) = IDF(t) × tf(t,d)×(k1+1) / [tf(t,d) + k1×(1-b + b×|d|/avgdl)]
        IDF(t)     = log( (N - df(t) + 0.5) / (df(t) + 0.5) + 1 )

    Why BM25 > TF-IDF for this dataset:
      k1=1.5 → TF saturation: word at 20x only scores ~3x more than once (not 20x)
      b=0.75 → Length norm: 300-word news article won't dominate 20-word tweet

    Role in pipeline: Stage-1 fast filter.
    Reduces 1774² = 3.1M pairs to ~88K candidates (2.8%) in one pass.
    Only those candidates are passed to the slower semantic stage.
    """
    def __init__(self, k1=1.5, b=0.75):
        self.k1, self.b = k1, b

    def fit(self, corpus: list):
        self.corpus   = corpus
        self.N        = len(corpus)
        self.doc_len  = [len(d) for d in corpus]
        self.avgdl    = sum(self.doc_len) / self.N if self.N else 1.0
        df_cnt        = Counter(t for doc in corpus for t in set(doc))
        self.idf      = {t: math.log((self.N-f+0.5)/(f+0.5)+1) for t,f in df_cnt.items()}
        self.tf_cache = [Counter(d) for d in corpus]
        return self

    def _score(self, qtoks, i) -> float:
        tf_map, dl = self.tf_cache[i], self.doc_len[i]
        s = 0.0
        for t in qtoks:
            if t not in self.idf: continue
            tf  = tf_map.get(t, 0)
            num = tf * (self.k1 + 1)
            den = tf + self.k1 * (1 - self.b + self.b * dl / self.avgdl)
            s  += self.idf[t] * (num / den)
        return s

    def get_candidates(self, top_k: int) -> list:
        """
        Return top_k candidate indices per document (excluding self).
        Scores are row-normalized to [0,1] for threshold comparison.
        """
        print(f"  BM25 scoring {self.N:,} docs (k1={self.k1}, b={self.b})...")
        results = []
        for i, doc in enumerate(self.corpus):
            scores     = np.array([self._score(doc, j) for j in range(self.N)], dtype=np.float32)
            scores[i]  = -1
            mx         = scores.max()
            if mx > 0: scores /= mx
            top        = np.argsort(scores)[::-1][:top_k]
            results.append((i, [(int(j), float(scores[j])) for j in top if scores[j] > 0]))
            if i % 300 == 0 and i > 0:
                print(f"    BM25: {i}/{self.N}...")
        total = sum(len(c) for _,c in results)
        print(f"  BM25 candidates: {total:,} pairs from {self.N}² = {self.N**2:,} possible "
              f"({total/self.N**2*100:.1f}% kept)")
        return results



# STAGE 2A — MULTILINGUAL SEMANTIC EMBEDDINGS (From Scratch)


def build_semantic_embeddings(df: pd.DataFrame) -> np.ndarray:
    """
    Multilingual semantic embeddings built from scratch using:
      Character trigrams (2-5 grams) → TF-IDF → TruncatedSVD (LSA)

    This is Latent Semantic Analysis (LSA) — the mathematical foundation
    that sentence transformers build upon with neural layers on top.

    Why character n-grams work for multilingual:
      - Devanagari script: trigrams capture subword morphology (prefixes/suffixes)
      - Hinglish: overlapping trigrams handle spelling variants
        "barish" and "baarish" share trigrams: "ari","ris","ish"
      - English: standard subword coverage

    Dimensions: 200 latent semantic dimensions (vs 384 for MiniLM)
    When sentence-transformers is installed, this is replaced by the neural model.
    """
    if _HAS_ST:
        print(f"\n[EMBED] Neural embeddings via {CFG['st_model']}...")
        model = SentenceTransformer(CFG["st_model"])
        embs  = model.encode(
            df["post_snippet"].tolist(),
            batch_size=CFG["st_batch"],
            show_progress_bar=True,
            normalize_embeddings=True,
            convert_to_numpy=True,
        )
        print(f"  Neural embeddings shape: {embs.shape}")
        return embs.astype(np.float32)

    print(f"\n[EMBED] Building multilingual LSA embeddings (from scratch)...")
    texts = df["post_snippet"].tolist()   # use original text, not preprocessed

    # Three complementary vectorizers for maximum multilingual coverage
    # 1. Character n-grams (2-5): captures subword patterns in ALL scripts
    cv1 = TfidfVectorizer(analyzer="char_wb", ngram_range=(2,5),
                          min_df=2, max_features=80000, sublinear_tf=True)
    # 2. Word unigrams + bigrams: captures phrase-level meaning
    wv  = TfidfVectorizer(analyzer="word", ngram_range=(1,2),
                          min_df=2, max_features=50000, sublinear_tf=True)
    # 3. Character n-grams on preprocessed text: cross-script overlap
    cv2 = TfidfVectorizer(analyzer="char_wb", ngram_range=(3,4),
                          min_df=2, max_features=50000, sublinear_tf=True)

    proc_texts = df["processed_text"].tolist()
    mat = hstack([
        cv1.fit_transform(texts)      * 0.5,   # char ngrams on original (most multilingual)
        wv.fit_transform(proc_texts)  * 0.3,   # word ngrams on cleaned text
        cv2.fit_transform(proc_texts) * 0.2,   # char ngrams on cleaned text
    ])

    vocab_total = len(cv1.vocabulary_) + len(wv.vocabulary_) + len(cv2.vocabulary_)
    print(f"  Combined vocab: {vocab_total:,} features across 3 vectorizers")

    # LSA: project sparse high-dim TF-IDF into dense semantic space
    # TruncatedSVD finds the 200 most important semantic dimensions
    dims = min(CFG["embed_dims"], mat.shape[1]-1, mat.shape[0]-1)
    svd  = TruncatedSVD(n_components=dims, n_iter=7, random_state=42)
    embs = normalize(svd.fit_transform(mat))   # L2 normalize → cosine = dot product

    explained = svd.explained_variance_ratio_.sum()
    print(f"  LSA embeddings: {embs.shape} | Variance explained: {explained:.1%}")
    return embs.astype(np.float32)



# STAGE 2B — ANN SEARCH (FAISS or sklearn NearestNeighbors)


class ANNIndex:
    """
    Approximate Nearest Neighbor index — uses FAISS if available,
    otherwise sklearn NearestNeighbors (exact cosine, SIMD-optimized).

    FAISS IndexFlatIP: exact inner product (= cosine on L2-normalized vectors)
    Advantage over sklearn cosine_similarity matrix:
      - sklearn: builds full N×N dense matrix → O(N²) memory
      - FAISS / NearestNeighbors: returns only top-k → O(N×k) memory
      - For N=1774: sklearn = 25MB OK; for N=100K: sklearn = 80GB → use ANN

    For 100K+ rows: switch FAISS to IndexIVFFlat (nlist=1000) for approximate
    search that is 50-100x faster with <1% accuracy loss.
    """
    def __init__(self, embeddings: np.ndarray):
        self.embeddings = embeddings
        self.N          = embeddings.shape[0]
        self.dim        = embeddings.shape[1]

        if _HAS_FAISS:
            print(f"  Building FAISS IndexFlatIP ({self.N} vectors, dim={self.dim})...")
            self.index = _faiss.IndexFlatIP(self.dim)
            self.index.add(embeddings)
            self._backend = "FAISS"
        else:
            print(f"  Building NearestNeighbors index ({self.N} vectors, dim={self.dim})...")
            self.index = NearestNeighbors(
                n_neighbors = min(self.N, CFG["top_n"] + 1),
                metric      = "cosine",
                algorithm   = "brute",   # exact for dense vectors
                n_jobs      = -1,        # use all CPU cores
            )
            self.index.fit(embeddings)
            self._backend = "sklearn-NearestNeighbors"
        print(f"  ANN backend: {self._backend}")

    def query(self, query_embs: np.ndarray, k: int) -> tuple:
        """Return (distances, indices) for top-k neighbors of each query."""
        if _HAS_FAISS:
            scores, indices = self.index.search(query_embs, k+1)  # +1 to exclude self
            # FAISS IP → convert to cosine distance for consistency
            distances = 1 - scores
        else:
            distances, indices = self.index.kneighbors(query_embs, n_neighbors=k+1)
        return distances, indices



# STAGE 2C — SEMANTIC RERANKING


def semantic_rerank(df, bm25_candidates, embeddings, ann_index, threshold, top_n) -> pd.DataFrame:
    """
    Rerank BM25 candidates using semantic similarity from embeddings.

    Two modes:
    1. Full ANN search (when FAISS available): query all embeddings at once,
       find globally top-k similar posts per post.
    2. Candidate-pair scoring: compute cosine only between each post and its
       50 BM25 candidates. Faster and more precise for candidate-based pipeline.

    Final score = semantic cosine similarity (0-1, higher = more similar).
    """
    print(f"\n[STAGE 2] Semantic reranking with {'neural' if _HAS_ST else 'LSA'} embeddings...")
    results = []

    # Use ANN for global top-k search (FAISS mode or when we want full coverage)
    if _HAS_FAISS:
        dists, indices = ann_index.query(embeddings, k=top_n)
        for i in range(len(df)):
            for k_idx, (dist, j) in enumerate(zip(dists[i], indices[i])):
                if j == i: continue
                score = float(1 - dist)   # cosine sim = 1 - cosine dist
                if score >= threshold:
                    results.append(_row(df, i, j, score))
    else:
        # Candidate-pair mode: score each post against its BM25 top-50
        for src_idx, candidates in bm25_candidates:
            if not candidates: continue
            cand_idx  = [c[0] for c in candidates]
            src_emb   = embeddings[src_idx].reshape(1, -1)
            cand_embs = embeddings[cand_idx]
            sims      = (src_emb @ cand_embs.T).flatten()  # dot = cosine (L2 normalized)
            order     = np.argsort(sims)[::-1][:top_n]
            for rank in order:
                score = float(sims[rank])
                if score >= threshold:
                    j = cand_idx[rank]
                    results.append(_row(df, src_idx, j, score))

    out = pd.DataFrame(results).drop_duplicates(subset=["source_id","matched_id"])
    print(f"  Semantic matches: {len(out):,} pairs (threshold≥{threshold})")
    return out

def _row(df, i, j, score):
    return {
        "source_id"            : int(df.iloc[i]["id"]),
        "source_username"      : df.iloc[i]["username"],
        "source_sheet"         : df.iloc[i]["sheet"],
        "source_post_snippet"  : df.iloc[i]["post_snippet"],
        "matched_id"           : int(df.iloc[j]["id"]),
        "matched_username"     : df.iloc[j]["username"],
        "matched_sheet"        : df.iloc[j]["sheet"],
        "matched_post_snippet" : df.iloc[j]["post_snippet"],
        "similarity_score"     : round(score, 4),
        "method"               : "neural-semantic" if _HAS_ST else "lsa-semantic",
    }



# TF-IDF COMPARISON


def compute_tfidf(df, top_n=5, threshold=0.15):
    """TF-IDF + cosine — included for method comparison sheet only."""
    print("\n[COMPARE] TF-IDF similarity (comparison method)...")
    texts = df["processed_text"].tolist()
    wv    = TfidfVectorizer(analyzer="word",    ngram_range=(1,2), min_df=2, max_features=50000, sublinear_tf=True)
    cv    = TfidfVectorizer(analyzer="char_wb", ngram_range=(2,4), min_df=2, max_features=50000, sublinear_tf=True)
    mat   = hstack([wv.fit_transform(texts)*0.6, cv.fit_transform(texts)*0.4])
    print(f"  Vocab: word={len(wv.vocabulary_):,}  char={len(cv.vocabulary_):,}")
    rows  = []
    for s in range(0, len(df), 200):
        e    = min(s+200, len(df))
        sims = cosine_similarity(mat[s:e], mat)
        for i, row in enumerate(sims):
            src = s+i; row[src] = -1
            for j in np.argsort(row)[::-1][:top_n]:
                sc = float(row[j])
                if sc >= threshold:
                    rows.append({"source_id":int(df.iloc[src]["id"]),
                                 "matched_id":int(df.iloc[j]["id"]),
                                 "tfidf_score":round(sc,4)})
    out = pd.DataFrame(rows)
    print(f"  TF-IDF matches: {len(out):,}")
    return out, mat



# FUZZY BASELINE


def compute_fuzzy(df, sample_size=150, top_n=3):
    """difflib SequenceMatcher baseline on a random sample."""
    print(f"\n[BASELINE] Fuzzy matching on {sample_size}-record sample...")
    sample = df.sample(min(sample_size, len(df)), random_state=42).reset_index(drop=True)
    texts  = sample["processed_text"].tolist()
    rows   = []
    for i in range(len(texts)):
        ranked = sorted(
            [(j, difflib.SequenceMatcher(None, texts[i], texts[j]).ratio())
             for j in range(len(texts)) if j!=i], key=lambda x:-x[1])
        for j, r in ranked[:top_n]:
            if r >= CFG["fuzzy_threshold"]:
                rows.append({"source_id":int(sample.iloc[i]["id"]),
                             "source_post_snippet":sample.iloc[i]["post_snippet"],
                             "matched_id":int(sample.iloc[j]["id"]),
                             "matched_post_snippet":sample.iloc[j]["post_snippet"],
                             "fuzzy_score":round(r,4)})
    out = pd.DataFrame(rows)
    print(f"  Fuzzy matches: {len(out)}")
    return out



# CLUSTERING ON SEMANTIC EMBEDDINGS


def cluster_posts(df, embeddings) -> pd.DataFrame:
    """
    Agglomerative clustering on L2-normalized semantic embeddings.
    Using cosine distance + average linkage.
    Cluster count = max(5, n // target_size).
    """
    print("\n[CLUSTER] Agglomerative clustering on semantic embeddings...")
    n_clusters = max(5, min(len(df)//CFG["cluster_size"], len(df)//3))
    print(f"  Target: {n_clusters} clusters ({CFG['cluster_size']} posts each) for {len(df):,} posts")
    labels = AgglomerativeClustering(
        n_clusters=n_clusters, metric="cosine", linkage="average"
    ).fit_predict(embeddings)
    df = df.copy()
    df["cluster_id"] = labels
    return df



# CLUSTER TITLE GENERATION — LLM or KEYWORD FALLBACK


def _call_claude_api(prompt: str, api_key: str, model: str) -> str:
    """
    Call Anthropic Claude API via raw HTTP (no SDK needed).
    Uses urllib from Python stdlib — works offline if API is reachable.
    """
    payload = json.dumps({
        "model"      : model,
        "max_tokens" : 40,
        "messages"   : [{"role": "user", "content": prompt}]
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data    = payload,
        method  = "POST",
        headers = {
            "Content-Type"      : "application/json",
            "x-api-key"         : api_key,
            "anthropic-version" : "2023-06-01",
        }
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return data["content"][0]["text"].strip().strip('"').strip("'")


def _keyword_title(texts: list) -> str:
    tokens = " ".join(texts).split()
    top    = [w for w,_ in Counter(tokens).most_common(20) if len(w)>2][:3]
    return " | ".join(top) if top else "Miscellaneous"


def generate_cluster_titles(df: pd.DataFrame) -> dict:
    """
    Generate human-readable topic titles per cluster.

    Priority:
      1. Claude API (if ANTHROPIC_API_KEY set and reachable)
         → Sends 3 sample posts, gets English topic title
      2. Keyword extraction fallback (always works)
         → Top-3 TF-weighted terms joined with |
    """
    api_key = CFG["anthropic_api_key"] or os.environ.get("ANTHROPIC_API_KEY", "")
    titles  = {}
    cids    = sorted(df["cluster_id"].unique())

    if api_key:
        print(f"\n[TITLES] Generating LLM titles via Claude API ({len(cids)} clusters)...")
        sizes   = df.groupby("cluster_id").size().sort_values(ascending=False)
        top_ids = set(sizes.index[:CFG["llm_max_clusters"]].tolist())
        success, fail = 0, 0

        for cid in cids:
            texts = df[df["cluster_id"]==cid]["processed_text"].tolist()
            if cid not in top_ids:
                titles[cid] = _keyword_title(texts)
                continue
            samples   = df[df["cluster_id"]==cid]["post_snippet"].tolist()[:3]
            posts_str = "\n---\n".join(s[:250] for s in samples)
            prompt    = (
                "These social media/news posts are from Uttar Pradesh, India and belong to the same topic cluster.\n"
                "Give ONE short topic title in English (5-8 words max).\n"
                "Reply with ONLY the title. No quotes. No explanation.\n\n"
                f"Posts:\n{posts_str}"
            )
            try:
                title = _call_claude_api(prompt, api_key, CFG["llm_model"])
                titles[cid] = title
                success += 1
                if success <= 5 or cid % 10 == 0:
                    print(f"  [{cid:3d}] '{title}'")
            except Exception as e:
                titles[cid] = _keyword_title(texts)
                fail += 1

        print(f"  LLM titles: {success} generated, {fail} used keyword fallback")
    else:
        print(f"\n[TITLES] Generating keyword titles ({len(cids)} clusters)...")
        print("  Tip: set ANTHROPIC_API_KEY environment variable for LLM-generated titles")
        for cid in cids:
            texts      = df[df["cluster_id"]==cid]["processed_text"].tolist()
            titles[cid] = _keyword_title(texts)

    return titles



# BUILD OUTPUT DATAFRAMES


def build_similarity_output(sim_df):
    cols = ["source_id","source_username","source_sheet","source_post_snippet",
            "matched_id","matched_username","matched_sheet","matched_post_snippet",
            "similarity_score","method"]
    return sim_df[[c for c in cols if c in sim_df.columns]]

def build_cluster_output(df, titles):
    df = df.copy()
    df["cluster_title"]    = df["cluster_id"].map(titles)
    umap = df.groupby("cluster_id")["username"].apply(
        lambda x: ", ".join(sorted(set(x.astype(str))))).to_dict()
    df["cluster_usernames"] = df["cluster_id"].map(umap)
    return df[["id","sheet","district","platform","username","post_snippet",
               "date","cluster_id","cluster_title","cluster_usernames"]]

def build_comparison(primary_df, tfidf_df, fuzzy_df, label):
    print("\n[COMPARE] Building 3-way method comparison table...")
    if primary_df.empty: return pd.DataFrame()
    comp = primary_df[["source_id","source_post_snippet","matched_id",
                        "matched_post_snippet","similarity_score","method"]].copy()
    comp = comp.rename(columns={"similarity_score":f"{label}_score","method":"pipeline"})
    if not tfidf_df.empty:
        comp = comp.merge(tfidf_df[["source_id","matched_id","tfidf_score"]],
                          on=["source_id","matched_id"], how="left")
        comp["tfidf_score"] = comp["tfidf_score"].fillna(0.0)
    else:
        comp["tfidf_score"] = 0.0
    if not fuzzy_df.empty:
        comp = comp.merge(fuzzy_df[["source_id","matched_id","fuzzy_score"]],
                          on=["source_id","matched_id"], how="left")
        comp["fuzzy_score"] = comp["fuzzy_score"].fillna("N/A")
    else:
        comp["fuzzy_score"] = "N/A"

    def note(row):
        try:
            s = float(row[f"{label}_score"]); t = float(row["tfidf_score"]); d = s-t
            if d > 0.15: return f"{label} +{d:.2f} — semantic understanding advantage"
            if d < -0.15: return f"TF-IDF +{-d:.2f} — shared rare keywords"
            return "All methods agree"
        except: return ""
    comp["analysis"] = comp.apply(note, axis=1)
    return comp



# SAVE OUTPUTS


def save_outputs(sim_out, cluster_out, cluster_sum, comparison, fuzzy_df,
                 output_path, pipeline_label, df):
    print(f"\n[SAVE] Writing results → {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as w:
        sim_out.to_excel(w,      sheet_name="Similarity_Matches",   index=False)
        cluster_out.to_excel(w,  sheet_name="Cluster_Assignments",  index=False)
        cluster_sum.to_excel(w,  sheet_name="Cluster_Summary",      index=False)
        if not comparison.empty:
            comparison.to_excel(w, sheet_name="Method_Comparison",  index=False)
        if not fuzzy_df.empty:
            fuzzy_df.to_excel(w, sheet_name="Fuzzy_Baseline",       index=False)

        # Pipeline info sheet
        pd.DataFrame([
            {"Setting": "Pipeline",             "Value": pipeline_label},
            {"Setting": "Embedding model",       "Value": CFG["st_model"] if _HAS_ST else "LSA (char ngrams + TruncatedSVD)"},
            {"Setting": "ANN backend",           "Value": "FAISS IndexFlatIP" if _HAS_FAISS else "sklearn NearestNeighbors (cosine)"},
            {"Setting": "LLM titles",            "Value": "Claude API" if (CFG["anthropic_api_key"] or os.environ.get("ANTHROPIC_API_KEY")) else "Keyword extraction"},
            {"Setting": "BM25 k1",               "Value": CFG["bm25_k1"]},
            {"Setting": "BM25 b",                "Value": CFG["bm25_b"]},
            {"Setting": "BM25 candidates / post","Value": CFG["bm25_candidates"]},
            {"Setting": "Semantic threshold",    "Value": CFG["semantic_threshold"]},
            {"Setting": "Total posts",           "Value": len(df)},
            {"Setting": "Total matches",         "Value": len(sim_out)},
            {"Setting": "Total clusters",        "Value": df["cluster_id"].nunique()},
            {"Setting": "Avg cluster size",      "Value": f"{len(df)/df['cluster_id'].nunique():.1f}"},
        ]).to_excel(w, sheet_name="Pipeline_Info", index=False)

    print(f"  Sheets written:")
    for s in ["Similarity_Matches","Cluster_Assignments","Cluster_Summary",
              "Method_Comparison","Fuzzy_Baseline","Pipeline_Info"]:
        print(f"    ✓ {s}")



# MAIN


def main():
    INPUT_FILE  = resolve_input_file()
    output_dir  = "/mnt/user-data/outputs" if os.path.isdir("/mnt/user-data/outputs") \
                  else os.path.dirname(os.path.abspath(INPUT_FILE))
    OUTPUT_FILE = os.path.join(output_dir, "post_similarity_results.xlsx")

    embed_label = "neural" if _HAS_ST else "lsa"
    ann_label   = "FAISS"  if _HAS_FAISS else "NearestNeighbors"
    llm_label   = "Claude API" if (CFG["anthropic_api_key"] or os.environ.get("ANTHROPIC_API_KEY")) else "keyword"

    pipeline_label = (
        f"BM25 (Stage1) → {'SentenceTransformer' if _HAS_ST else 'LSA-Semantic'} (Stage2)"
        f" | ANN: {ann_label} | Titles: {llm_label}"
    )

    print("=" * 66)
    print(" POST SNIPPET SIMILARITY MATCHER — FULL PIPELINE")
    print(f"  Embeddings : {'SentenceTransformer (' + CFG['st_model'] + ')' if _HAS_ST else 'LSA (char ngrams + TruncatedSVD 200-dim)'}")
    print(f"  ANN Search : {ann_label}")
    print(f"  LLM Titles : {llm_label}")
    print(f"  Languages  : Hindi (Devanagari) | English | Hinglish (Roman)")
    print("=" * 66)

    # Load & preprocess
    df              = load_all_sheets(INPUT_FILE)
    df              = preprocess_df(df)

    # Stage 1: BM25 candidate filtering
    print("\n[STAGE 1] BM25 fast candidate filtering...")
    tokenized       = [t.split() for t in df["processed_text"].tolist()]
    bm25            = BM25(k1=CFG["bm25_k1"], b=CFG["bm25_b"])
    bm25.fit(tokenized)
    print(f"  Vocab: {len(bm25.idf):,} terms | Avg doc: {bm25.avgdl:.1f} tokens")
    candidates      = bm25.get_candidates(top_k=CFG["bm25_candidates"])

    # Stage 2: Semantic embeddings + ANN reranking
    embeddings      = build_semantic_embeddings(df)
    ann_index       = ANNIndex(embeddings)
    primary_df      = semantic_rerank(df, candidates, embeddings, ann_index,
                                      CFG["semantic_threshold"], CFG["top_n"])

    # Comparison methods
    tfidf_df, tfidf_mat = compute_tfidf(df, CFG["top_n"], CFG["tfidf_threshold"])
    fuzzy_df        = compute_fuzzy(df, CFG["fuzzy_sample"])

    # Clustering on semantic embeddings
    df              = cluster_posts(df, embeddings)

    # Cluster titles
    titles          = generate_cluster_titles(df)

    # Build outputs
    sim_out         = build_similarity_output(primary_df)
    cluster_out     = build_cluster_output(df, titles)
    cluster_sum     = cluster_out.groupby(["cluster_id","cluster_title"]).agg(
        num_posts   = ("id",       "count"),
        usernames   = ("username", lambda x: ", ".join(sorted(set(x.astype(str))))),
        platforms   = ("platform", lambda x: ", ".join(sorted(set(x.astype(str))))),
        sample_post = ("post_snippet", "first"),
    ).reset_index()
    comparison      = build_comparison(primary_df, tfidf_df, fuzzy_df, embed_label)

    # Save
    os.makedirs(output_dir, exist_ok=True)
    save_outputs(sim_out, cluster_out, cluster_sum, comparison, fuzzy_df,
                 OUTPUT_FILE, pipeline_label, df)

    print("\n" + "=" * 66)
    print(" SUMMARY")
    print("=" * 66)
    print(f"  Pipeline          : {pipeline_label}")
    print(f"  Posts processed   : {len(df):,}")
    print(f"  Similarity matches: {len(sim_out):,}")
    print(f"  TF-IDF matches    : {len(tfidf_df):,}  (comparison)")
    print(f"  Clusters          : {df['cluster_id'].nunique()}")
    print(f"  Avg cluster size  : {len(df)/df['cluster_id'].nunique():.1f} posts")
    if not sim_out.empty:
        print(f"  Avg score         : {sim_out['similarity_score'].mean():.4f}")
        print(f"  Max score         : {sim_out['similarity_score'].max():.4f}")
    print(f"\n  Output: {OUTPUT_FILE}")
    print("=" * 66)
    print("\n  Done!")

    if not _HAS_ST:
        print("\n  To unlock neural multilingual embeddings:")
        print("    pip install sentence-transformers")
    if not _HAS_FAISS:
        print("  To unlock FAISS ANN index:")
        print("    pip install faiss-cpu")
    if not (CFG["anthropic_api_key"] or os.environ.get("ANTHROPIC_API_KEY")):
        print("  To unlock LLM cluster titles:")
        print("    export ANTHROPIC_API_KEY=sk-ant-...")


if __name__ == "__main__":
    main()
